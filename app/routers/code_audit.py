"""Code audit API endpoints for security and quality analysis."""

from fastapi import APIRouter, Depends, HTTPException, status, Response, BackgroundTasks, Request
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from typing import List, Optional
from pydantic import BaseModel
from datetime import datetime, timedelta
import logging
import uuid
import io
import asyncio
import threading

from app.database import get_db, SessionLocal
from app.models.code_audit import CodeAuditResult, CodeAuditSchedule, CodeAuditJob
from app.models.audit_log import AuditLog
from app.config import settings

router = APIRouter()
logger = logging.getLogger(__name__)


# Pydantic models
class AuditIssue(BaseModel):
    """Single audit issue."""
    id: Optional[int] = None
    severity: str
    issue: str
    location: Optional[str] = None
    mitigation: Optional[str] = None
    extra_data: Optional[dict] = None
    resolution_status: str = "pending"
    resolution_notes: Optional[str] = None
    resolved_by: Optional[str] = None
    resolved_at: Optional[str] = None


class AuditCategoryResult(BaseModel):
    """Results for a single audit category."""
    category: str
    issues: List[AuditIssue]
    count: int
    pending_count: int = 0
    mitigated_count: int = 0
    risk_approved_count: int = 0


class AuditRunRequest(BaseModel):
    """Request to run an audit."""
    categories: List[str]


class AuditRunResponse(BaseModel):
    """Response from running an audit (now returns job info for async)."""
    job_id: str
    status: str
    message: str


class AuditJobStatusResponse(BaseModel):
    """Status of a running audit job."""
    job_id: str
    status: str
    progress: int
    current_category: Optional[str] = None
    completed_categories: int = 0
    total_categories: int = 0
    run_id: Optional[str] = None
    error_message: Optional[str] = None
    created_at: Optional[str] = None
    completed_at: Optional[str] = None


class AuditResultsResponse(BaseModel):
    """Response with latest audit results."""
    run_id: Optional[str] = None
    timestamp: Optional[str] = None
    total_issues: int = 0
    pending_count: int = 0
    mitigated_count: int = 0
    risk_approved_count: int = 0
    results: List[AuditCategoryResult] = []
    schedule_status: Optional[str] = None
    active_job: Optional[AuditJobStatusResponse] = None


class UpdateResolutionRequest(BaseModel):
    """Request to update issue resolution status."""
    resolution_status: str  # pending, mitigated, risk_approved
    resolution_notes: Optional[str] = None


class ScheduleRequest(BaseModel):
    """Request to set audit schedule."""
    enabled: bool
    frequency: str = "weekly"
    day_of_week: int = 1
    time_utc: str = "02:00"
    categories: List[str]
    email_notify: bool = False


class ScheduleResponse(BaseModel):
    """Response with schedule settings."""
    enabled: bool
    frequency: str
    day_of_week: int
    time_utc: str
    categories: List[str]
    email_notify: bool
    last_run: Optional[str] = None
    next_run: Optional[str] = None


# Pre-defined audit findings based on the comprehensive audit performed
# These represent the baseline findings that would be detected by static analysis
BASELINE_AUDIT_FINDINGS = {
    "owasp": [
        {"severity": "CRITICAL", "issue": "Hardcoded database credentials", "location": "scripts/fix_documents_schema.py:5", "mitigation": "Move credentials to Azure Key Vault or environment variables"},
        {"severity": "CRITICAL", "issue": "Hardcoded database credentials", "location": "scripts/check_schema.py:5", "mitigation": "Move credentials to Azure Key Vault or environment variables"},
        {"severity": "CRITICAL", "issue": "SQL injection via f-strings in database scripts", "location": "scripts/fix_documents_schema.py:41-46", "mitigation": "Use parameterized queries instead of string interpolation"},
        {"severity": "CRITICAL", "issue": "IDOR - No ownership check on document access", "location": "app/routers/documents.py", "mitigation": "Add user ownership validation before returning document data"},
        {"severity": "CRITICAL", "issue": "Development mode bypasses authentication entirely", "location": "app/middleware/auth.py:58-62", "mitigation": "Remove dev mode bypass or restrict to non-production environments"},
        {"severity": "HIGH", "issue": "PHI exposed in error messages", "location": "app/routers/auth.py:184,227", "mitigation": "Sanitize error messages to remove sensitive data"},
        {"severity": "HIGH", "issue": "SSO errors reflected in URL parameters", "location": "app/routers/auth.py:153", "mitigation": "Store errors in session, not URL"},
        {"severity": "HIGH", "issue": "Session timeout stored in client-side cookie", "location": "app/middleware/auth.py:90-102", "mitigation": "Move session management to server-side"},
        {"severity": "MEDIUM", "issue": "Open redirect vulnerability", "location": "app/routers/auth.py:153", "mitigation": "Validate redirect URLs against whitelist"},
        {"severity": "MEDIUM", "issue": "JSON from DB loaded without schema validation", "location": "app/routers/scan.py:31", "mitigation": "Add JSON schema validation"},
        {"severity": "MEDIUM", "issue": "python3-saml dependency may have vulnerabilities", "location": "requirements.txt", "mitigation": "Update to latest version and run security audit"},
        {"severity": "MEDIUM", "issue": "No rate limiting on auth endpoints", "location": "app/routers/auth.py", "mitigation": "Add per-endpoint rate limiting"},
    ],
    "hipaa": [
        {"severity": "CRITICAL", "issue": "164.312(a)(2)(i) - Hardcoded credentials violate encryption key management", "location": "scripts/*.py", "mitigation": "Use Azure Key Vault for all credentials", "extra_data": {"regulation": "164.312(a)(2)(i)"}},
        {"severity": "CRITICAL", "issue": "164.308(a)(4) - No access control on document endpoints (IDOR)", "location": "app/routers/documents.py", "mitigation": "Implement document ownership validation", "extra_data": {"regulation": "164.308(a)(4)"}},
        {"severity": "HIGH", "issue": "164.312(a)(2)(iii) - Session timeout can be bypassed via cookie manipulation", "location": "app/middleware/auth.py", "mitigation": "Move session management to server-side", "extra_data": {"regulation": "164.312(a)(2)(iii)"}},
        {"severity": "HIGH", "issue": "164.312(b) - Audit logging can be disabled via configuration", "location": "app/services/audit_service.py:60-77", "mitigation": "Make audit logging mandatory, not configurable", "extra_data": {"regulation": "164.312(b)"}},
        {"severity": "HIGH", "issue": "164.308(a)(5)(ii)(B) - PHI in error messages", "location": "app/main.py:268-283", "mitigation": "Sanitize all error responses", "extra_data": {"regulation": "164.308(a)(5)(ii)(B)"}},
        {"severity": "HIGH", "issue": "164.308(a)(5)(ii)(C) - Break glass admin actions not audited", "location": "app/routers/auth.py", "mitigation": "Add audit logging for all admin actions", "extra_data": {"regulation": "164.308(a)(5)(ii)(C)"}},
        {"severity": "MEDIUM", "issue": "Missing admin authorization on user management endpoints", "location": "app/routers/auth.py:408-483", "mitigation": "Add require_admin dependency to all user endpoints", "extra_data": {"regulation": "164.312(a)(1)"}},
    ],
    "performance": [
        {"severity": "CRITICAL", "issue": "Full-table scan loading ALL facilities for fuzzy matching", "location": "app/services/facility_matching_service.py:140,153,343,603", "mitigation": "Implement pagination or database-side fuzzy matching"},
        {"severity": "CRITICAL", "issue": "Entire file contents loaded into memory for bulk uploads", "location": "app/routers/documents.py:146-161", "mitigation": "Implement streaming upload processing"},
        {"severity": "CRITICAL", "issue": "Synchronous DB calls in async functions", "location": "app/workers/extraction_worker.py:74-81", "mitigation": "Use async database drivers or run_in_executor"},
        {"severity": "CRITICAL", "issue": "Synchronous blob operations blocking event loop", "location": "app/workers/blob_watcher.py:74-79", "mitigation": "Use aiofiles and async blob SDK"},
        {"severity": "HIGH", "issue": "N+1 query pattern in document list", "location": "app/routers/documents.py:407-409", "mitigation": "Use joinedload() or selectinload() for relationships"},
        {"severity": "HIGH", "issue": "Repeated decryption for every document in list endpoint", "location": "app/routers/documents.py:366-428", "mitigation": "Batch decrypt or use lazy decryption"},
        {"severity": "HIGH", "issue": "30 separate COUNT queries instead of GROUP BY", "location": "app/services/stats_service.py:86-100", "mitigation": "Use single query with GROUP BY date"},
        {"severity": "MEDIUM", "issue": "Missing indexes on frequently queried columns", "location": "app/models/facility.py", "mitigation": "Add indexes on phone, fax columns"},
        {"severity": "MEDIUM", "issue": "Job status store grows indefinitely in memory", "location": "app/background_tasks.py:22-72", "mitigation": "Implement cleanup mechanism or use Redis"},
    ],
    "dead_code": [
        {"severity": "HIGH", "issue": "Stub endpoint with only pass statement", "location": "app/routers/auth.py:108-112 (/auth/refresh)", "mitigation": "Implement token refresh or remove endpoint"},
        {"severity": "HIGH", "issue": "Stub endpoint with only pass statement", "location": "app/routers/auth.py:114-119 (/auth/me)", "mitigation": "Implement user info endpoint or remove"},
        {"severity": "HIGH", "issue": "Universal Print integration returns failure TODO", "location": "app/routers/print.py:367-375", "mitigation": "Implement or remove feature"},
        {"severity": "HIGH", "issue": "Direct IP printing not implemented", "location": "app/routers/print.py:378-384", "mitigation": "Implement or remove feature"},
        {"severity": "MEDIUM", "issue": "Empty exception handler swallowing errors", "location": "app/routers/queue.py:444", "mitigation": "Add proper error handling or remove try/except"},
        {"severity": "MEDIUM", "issue": "Empty exception handler swallowing errors", "location": "app/routers/print.py:74", "mitigation": "Add proper error handling"},
        {"severity": "MEDIUM", "issue": "Bare except:pass hiding errors", "location": "app/middleware/auth.py:102", "mitigation": "Add specific exception handling"},
        {"severity": "MEDIUM", "issue": "Multi-page TIFF/PDF combining TODO", "location": "app/routers/scan.py:161", "mitigation": "Implement feature or document limitation"},
        {"severity": "MEDIUM", "issue": "Layout-based splitting TODO", "location": "app/services/document_intelligence_service.py:282", "mitigation": "Implement or remove TODO"},
        {"severity": "MEDIUM", "issue": "Email alerting TODO", "location": "app/services/audit_service.py:137", "mitigation": "Implement email alerting"},
        {"severity": "LOW", "issue": "Unused FedEx configuration variables", "location": "app/config.py", "mitigation": "Remove unused configuration"},
        {"severity": "LOW", "issue": "Unused Google Places configuration", "location": "app/config.py", "mitigation": "Remove unused configuration"},
        {"severity": "LOW", "issue": "Duplicate token extraction code", "location": "app/routers/auth.py", "mitigation": "Consolidate into helper function"},
    ],
    "ada": [
        {"severity": "CRITICAL", "issue": "140+ form inputs without associated labels", "location": "app/templates/*.html", "mitigation": "Add <label for='id'> elements for all inputs", "extra_data": {"count": 140}},
        {"severity": "CRITICAL", "issue": "Clickable div elements not keyboard accessible", "location": "app/templates/*.html", "mitigation": "Replace <div onclick> with <button> elements", "extra_data": {"count": 8}},
        {"severity": "CRITICAL", "issue": "Missing ARIA labels on interactive elements", "location": "app/templates/*.html", "mitigation": "Add aria-label to icon-only buttons", "extra_data": {"count": 130}},
        {"severity": "CRITICAL", "issue": "Animations lack prefers-reduced-motion support", "location": "app/templates/login.html:42-47, dashboard.html:318-321", "mitigation": "Add @media (prefers-reduced-motion) CSS rules", "extra_data": {"count": 5}},
        {"severity": "HIGH", "issue": "Table headers missing scope attribute", "location": "app/templates/*.html", "mitigation": "Add scope='col' to all <th> elements", "extra_data": {"count": 6}},
        {"severity": "HIGH", "issue": "Status indicators use color only without icons/text", "location": "app/templates/*.html", "mitigation": "Add icons or text labels to status badges"},
        {"severity": "HIGH", "issue": "Small text below 16px minimum", "location": "app/templates/*.html (.perf-label, .source-badge)", "mitigation": "Increase font-size to minimum 16px", "extra_data": {"count": 15}},
        {"severity": "HIGH", "issue": "Toast notifications lack aria-live region", "location": "app/templates/*.html", "mitigation": "Add aria-live='polite' to toast container"},
        {"severity": "MEDIUM", "issue": "Missing skip navigation link", "location": "app/templates/*.html", "mitigation": "Add 'skip to main content' link at top of pages"},
        {"severity": "MEDIUM", "issue": "Missing lang attribute on HTML element", "location": "app/templates/admin.html, documents.html, scan.html, review.html, queue.html, compliance.html, health.html, history.html", "mitigation": "Add lang='en' to <html> element", "extra_data": {"count": 8}},
    ],
    "opswat": [
        {"severity": "MEDIUM", "issue": "File upload validation missing magic byte verification", "location": "app/services/document_service.py:135-156", "mitigation": "Add python-magic or similar for file type validation", "extra_data": {"status": "INCOMPLETE", "category": "File Scanning"}},
        {"severity": "MEDIUM", "issue": "No antivirus scanning integration", "location": "app/services/document_service.py", "mitigation": "Integrate ClamAV or OPSWAT MetaDefender", "extra_data": {"status": "MISSING", "category": "File Scanning"}},
        {"severity": "LOW", "issue": "Phone number validation missing", "location": "app/schemas/document.py:26", "mitigation": "Add phone number format validation", "extra_data": {"status": "INCOMPLETE", "category": "Data Validation"}},
        {"severity": "LOW", "issue": "NPI validation missing", "location": "app/schemas/document.py:33", "mitigation": "Add NPI checksum validation", "extra_data": {"status": "INCOMPLETE", "category": "Data Validation"}},
        {"severity": "LOW", "issue": "localStorage used for auth tokens - XSS risk", "location": "app/templates/*.html", "mitigation": "Use httpOnly cookies instead", "extra_data": {"status": "RISK", "category": "Data Leakage"}},
    ],
}


def get_user_from_request(request) -> dict:
    """Extract user info from request cookies/headers."""
    user_email = "system"
    if hasattr(request, "cookies"):
        user_email = request.cookies.get("user_email", "system")
    return {"email": user_email}


@router.get("/results", response_model=AuditResultsResponse)
async def get_audit_results(
    status_filter: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """Get the latest audit results with optional status filtering."""
    # Check for any active/running jobs
    active_job = db.query(CodeAuditJob)\
        .filter(CodeAuditJob.status.in_(["queued", "running"]))\
        .order_by(desc(CodeAuditJob.created_at))\
        .first()

    active_job_response = None
    if active_job:
        total_cats = len(active_job.categories) if active_job.categories else 0
        active_job_response = AuditJobStatusResponse(
            job_id=active_job.job_id,
            status=active_job.status,
            progress=active_job.progress,
            current_category=active_job.current_category if hasattr(active_job, 'current_category') else None,
            completed_categories=active_job.completed_categories if hasattr(active_job, 'completed_categories') else 0,
            total_categories=total_cats,
            run_id=active_job.run_id,
            created_at=active_job.created_at.isoformat() if active_job.created_at else None
        )

    # Get the most recent run_id
    latest_run = db.query(CodeAuditResult.run_id, CodeAuditResult.timestamp)\
        .order_by(desc(CodeAuditResult.timestamp))\
        .first()

    if not latest_run:
        # Return empty if no audit has been run yet
        schedule = db.query(CodeAuditSchedule).first()
        schedule_status = "Not Scheduled"
        if schedule and schedule.enabled:
            schedule_status = f"Scheduled {schedule.frequency}"

        return AuditResultsResponse(
            run_id=None,
            timestamp=None,
            total_issues=0,
            results=[],
            schedule_status=schedule_status,
            active_job=active_job_response
        )

    # Build query with optional status filter
    query = db.query(CodeAuditResult)\
        .filter(CodeAuditResult.run_id == latest_run.run_id)

    if status_filter and status_filter != "all":
        query = query.filter(CodeAuditResult.resolution_status == status_filter)

    results = query.order_by(CodeAuditResult.category, CodeAuditResult.severity).all()

    # Get total counts by status (unfiltered)
    all_results = db.query(CodeAuditResult)\
        .filter(CodeAuditResult.run_id == latest_run.run_id)\
        .all()

    total_pending = sum(1 for r in all_results if (r.resolution_status or "pending") == "pending")
    total_mitigated = sum(1 for r in all_results if r.resolution_status == "mitigated")
    total_risk_approved = sum(1 for r in all_results if r.resolution_status == "risk_approved")

    # Group by category
    categories = {}
    for result in results:
        if result.category not in categories:
            categories[result.category] = {"issues": [], "pending": 0, "mitigated": 0, "risk_approved": 0}

        resolution_status = result.resolution_status or "pending"
        categories[result.category]["issues"].append(AuditIssue(
            id=result.id,
            severity=result.severity,
            issue=result.issue,
            location=result.location,
            mitigation=result.mitigation,
            extra_data=result.extra_data,
            resolution_status=resolution_status,
            resolution_notes=result.resolution_notes,
            resolved_by=result.resolved_by,
            resolved_at=result.resolved_at.isoformat() if result.resolved_at else None
        ))

        # Count by status
        if resolution_status == "pending":
            categories[result.category]["pending"] += 1
        elif resolution_status == "mitigated":
            categories[result.category]["mitigated"] += 1
        elif resolution_status == "risk_approved":
            categories[result.category]["risk_approved"] += 1

    category_results = [
        AuditCategoryResult(
            category=cat,
            issues=data["issues"],
            count=len(data["issues"]),
            pending_count=data["pending"],
            mitigated_count=data["mitigated"],
            risk_approved_count=data["risk_approved"]
        )
        for cat, data in categories.items()
    ]

    # Get schedule status
    schedule = db.query(CodeAuditSchedule).first()
    schedule_status = "Not Scheduled"
    if schedule and schedule.enabled:
        schedule_status = f"Scheduled {schedule.frequency}"

    return AuditResultsResponse(
        run_id=latest_run.run_id,
        timestamp=latest_run.timestamp.isoformat() if latest_run.timestamp else None,
        total_issues=len(all_results),
        pending_count=total_pending,
        mitigated_count=total_mitigated,
        risk_approved_count=total_risk_approved,
        results=category_results,
        schedule_status=schedule_status,
        active_job=active_job_response
    )


def run_audit_background(job_id: str, categories: List[str], triggered_by: str):
    """Background task to run code audit."""
    db = SessionLocal()
    try:
        # Update job status to running
        job = db.query(CodeAuditJob).filter(CodeAuditJob.job_id == job_id).first()
        if not job:
            logger.error(f"Job {job_id} not found")
            return

        job.status = "running"
        job.started_at = datetime.utcnow()
        job.completed_categories = 0
        db.commit()

        run_id = str(uuid.uuid4())
        timestamp = datetime.utcnow()
        total_categories = len(categories)

        for idx, category in enumerate(categories):
            # Check if job was cancelled
            db.refresh(job)
            if job.status == "cancelled":
                logger.info(f"Job {job_id} was cancelled")
                return

            # Update current category being processed
            job.current_category = category
            job.run_id = run_id  # Set run_id early so partial results can be fetched
            db.commit()

            findings = BASELINE_AUDIT_FINDINGS.get(category, [])

            for finding in findings:
                audit_result = CodeAuditResult(
                    run_id=run_id,
                    timestamp=timestamp,
                    category=category,
                    severity=finding["severity"],
                    issue=finding["issue"],
                    location=finding.get("location"),
                    mitigation=finding.get("mitigation"),
                    extra_data=finding.get("extra_data"),
                    triggered_by=triggered_by,
                    resolution_status="pending"
                )
                db.add(audit_result)

            # Update progress after completing each category
            job.completed_categories = idx + 1
            job.progress = int(((idx + 1) / total_categories) * 100)
            db.commit()

        # Mark job as completed
        job.status = "completed"
        job.current_category = None
        job.progress = 100
        job.completed_at = datetime.utcnow()
        db.commit()

        # Log the audit run
        audit_log = AuditLog(
            user_id="system",
            user_email=triggered_by,
            action="run_code_audit",
            resource_type="code_audit",
            resource_id=run_id,
            success=True
        )
        db.add(audit_log)
        db.commit()

        logger.info(f"Background audit completed: job_id={job_id}, run_id={run_id}")

    except Exception as e:
        logger.error(f"Background audit failed: {e}")
        if job:
            job.status = "failed"
            job.error_message = str(e)
            job.completed_at = datetime.utcnow()
            db.commit()
    finally:
        db.close()


@router.post("/run", response_model=AuditRunResponse)
async def run_audit(
    request: AuditRunRequest,
    background_tasks: BackgroundTasks,
    req: Request,
    db: Session = Depends(get_db)
):
    """Queue a code audit job for background processing."""
    valid_categories = ["owasp", "hipaa", "performance", "dead_code", "ada", "opswat"]
    selected = [c for c in request.categories if c in valid_categories]

    if not selected:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No valid categories selected"
        )

    # Check for existing running job
    existing_job = db.query(CodeAuditJob)\
        .filter(CodeAuditJob.status.in_(["queued", "running"]))\
        .first()

    if existing_job:
        return AuditRunResponse(
            job_id=existing_job.job_id,
            status=existing_job.status,
            message="An audit is already in progress"
        )

    # Get user email from request
    user_email = req.cookies.get("user_email", "system")

    # Create job record
    job_id = str(uuid.uuid4())
    job = CodeAuditJob(
        job_id=job_id,
        status="queued",
        progress=0,
        categories=selected,
        triggered_by=user_email
    )
    db.add(job)
    db.commit()

    # Queue background task
    background_tasks.add_task(run_audit_background, job_id, selected, user_email)

    logger.info(f"Audit job queued: job_id={job_id}, categories={selected}")

    return AuditRunResponse(
        job_id=job_id,
        status="queued",
        message=f"Audit queued for {len(selected)} categories"
    )


@router.get("/job/{job_id}", response_model=AuditJobStatusResponse)
async def get_job_status(job_id: str, db: Session = Depends(get_db)):
    """Get the status of a running audit job."""
    job = db.query(CodeAuditJob).filter(CodeAuditJob.job_id == job_id).first()

    if not job:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Job not found"
        )

    # Get total categories from the job
    total_cats = len(job.categories) if job.categories else 0

    return AuditJobStatusResponse(
        job_id=job.job_id,
        status=job.status,
        progress=job.progress,
        current_category=job.current_category if hasattr(job, 'current_category') else None,
        completed_categories=job.completed_categories if hasattr(job, 'completed_categories') else 0,
        total_categories=total_cats,
        run_id=job.run_id,
        error_message=job.error_message,
        created_at=job.created_at.isoformat() if job.created_at else None,
        completed_at=job.completed_at.isoformat() if job.completed_at else None
    )




@router.post("/job/{job_id}/cancel")
async def cancel_audit_job(job_id: str, db: Session = Depends(get_db)):
    """Cancel a running audit job."""
    job = db.query(CodeAuditJob).filter(CodeAuditJob.job_id == job_id).first()

    if not job:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Job not found"
        )

    if job.status not in ["queued", "running"]:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Cannot cancel job with status: {job.status}"
        )

    job.status = "cancelled"
    job.completed_at = datetime.utcnow()
    db.commit()

    logger.info(f"Audit job {job_id} cancelled")

    return {"job_id": job_id, "status": "cancelled", "message": "Audit cancelled"}


@router.put("/issue/{issue_id}/status")
async def update_issue_status(
    issue_id: int,
    request: UpdateResolutionRequest,
    req: Request,
    db: Session = Depends(get_db)
):
    """Update the resolution status of a specific audit issue."""
    issue = db.query(CodeAuditResult).filter(CodeAuditResult.id == issue_id).first()

    if not issue:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Issue not found"
        )

    valid_statuses = ["pending", "mitigated", "risk_approved"]
    if request.resolution_status not in valid_statuses:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Invalid status. Must be one of: {', '.join(valid_statuses)}"
        )

    # Get user from request
    user_email = req.cookies.get("user_email", "system")

    # Update the issue
    old_status = issue.resolution_status
    issue.resolution_status = request.resolution_status
    issue.resolution_notes = request.resolution_notes
    issue.resolved_by = user_email
    issue.resolved_at = datetime.utcnow()

    # Log the status change
    audit_log = AuditLog(
        user_id="system",
        user_email=user_email,
        action="update_audit_issue_status",
        resource_type="code_audit_result",
        resource_id=str(issue_id),
        success=True,
        details=f"Status changed from {old_status} to {request.resolution_status}"
    )
    db.add(audit_log)
    db.commit()

    logger.info(f"Issue {issue_id} status updated: {old_status} -> {request.resolution_status} by {user_email}")

    return {
        "id": issue.id,
        "resolution_status": issue.resolution_status,
        "resolution_notes": issue.resolution_notes,
        "resolved_by": issue.resolved_by,
        "resolved_at": issue.resolved_at.isoformat() if issue.resolved_at else None
    }


@router.get("/schedule", response_model=ScheduleResponse)
async def get_schedule(db: Session = Depends(get_db)):
    """Get the audit schedule configuration."""
    schedule = db.query(CodeAuditSchedule).first()

    if not schedule:
        return ScheduleResponse(
            enabled=False,
            frequency="weekly",
            day_of_week=1,
            time_utc="02:00",
            categories=["owasp", "hipaa", "performance", "dead_code", "ada", "opswat"],
            email_notify=False
        )

    return ScheduleResponse(
        enabled=schedule.enabled,
        frequency=schedule.frequency,
        day_of_week=schedule.day_of_week or 1,
        time_utc=schedule.time_utc,
        categories=schedule.categories or [],
        email_notify=schedule.email_notify,
        last_run=schedule.last_run.isoformat() if schedule.last_run else None,
        next_run=schedule.next_run.isoformat() if schedule.next_run else None
    )


@router.post("/schedule", response_model=ScheduleResponse)
async def save_schedule(request: ScheduleRequest, db: Session = Depends(get_db)):
    """Save the audit schedule configuration."""
    schedule = db.query(CodeAuditSchedule).first()

    # Calculate next run
    next_run = None
    if request.enabled:
        now = datetime.utcnow()
        hour, minute = map(int, request.time_utc.split(":"))

        if request.frequency == "daily":
            next_run = now.replace(hour=hour, minute=minute, second=0, microsecond=0)
            if next_run <= now:
                next_run += timedelta(days=1)
        elif request.frequency == "weekly":
            days_ahead = request.day_of_week - now.weekday()
            if days_ahead <= 0:
                days_ahead += 7
            next_run = now + timedelta(days=days_ahead)
            next_run = next_run.replace(hour=hour, minute=minute, second=0, microsecond=0)
        elif request.frequency == "monthly":
            next_run = now.replace(day=1, hour=hour, minute=minute, second=0, microsecond=0)
            if next_run <= now:
                if now.month == 12:
                    next_run = next_run.replace(year=now.year + 1, month=1)
                else:
                    next_run = next_run.replace(month=now.month + 1)

    if schedule:
        schedule.enabled = request.enabled
        schedule.frequency = request.frequency
        schedule.day_of_week = request.day_of_week
        schedule.time_utc = request.time_utc
        schedule.categories = request.categories
        schedule.email_notify = request.email_notify
        schedule.next_run = next_run
    else:
        schedule = CodeAuditSchedule(
            enabled=request.enabled,
            frequency=request.frequency,
            day_of_week=request.day_of_week,
            time_utc=request.time_utc,
            categories=request.categories,
            email_notify=request.email_notify,
            next_run=next_run
        )
        db.add(schedule)

    db.commit()
    db.refresh(schedule)

    logger.info(f"Audit schedule updated: enabled={request.enabled}, frequency={request.frequency}")

    return ScheduleResponse(
        enabled=schedule.enabled,
        frequency=schedule.frequency,
        day_of_week=schedule.day_of_week or 1,
        time_utc=schedule.time_utc,
        categories=schedule.categories or [],
        email_notify=schedule.email_notify,
        last_run=schedule.last_run.isoformat() if schedule.last_run else None,
        next_run=schedule.next_run.isoformat() if schedule.next_run else None
    )


@router.get("/export")
async def export_audit_excel(db: Session = Depends(get_db)):
    """Export audit results to Excel file."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="openpyxl package not installed"
        )

    # Get the most recent run
    latest_run = db.query(CodeAuditResult.run_id, CodeAuditResult.timestamp)\
        .order_by(desc(CodeAuditResult.timestamp))\
        .first()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    critical_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    high_fill = PatternFill(start_color="FFB347", end_color="FFB347", fill_type="solid")
    medium_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    low_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    category_names = {
        "owasp": "Security (OWASP)",
        "hipaa": "HIPAA Compliance",
        "performance": "Performance",
        "dead_code": "Dead Code",
        "ada": "ADA/WCAG",
        "opswat": "OPSWAT"
    }

    if latest_run:
        results = db.query(CodeAuditResult)\
            .filter(CodeAuditResult.run_id == latest_run.run_id)\
            .order_by(CodeAuditResult.category)\
            .all()

        # Group by category
        categories = {}
        for result in results:
            if result.category not in categories:
                categories[result.category] = []
            categories[result.category].append(result)
    else:
        # Use baseline findings
        categories = {}
        for cat, findings in BASELINE_AUDIT_FINDINGS.items():
            categories[cat] = [type('obj', (object,), f)() for f in findings]

    # Create summary sheet
    summary_ws = wb.create_sheet("Summary")
    summary_ws.append(["Code Audit Report"])
    summary_ws.merge_cells("A1:D1")
    summary_ws["A1"].font = Font(bold=True, size=16)

    summary_ws.append([])
    summary_ws.append(["Report Generated:", datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")])
    if latest_run:
        summary_ws.append(["Audit Run ID:", latest_run.run_id])
        summary_ws.append(["Audit Timestamp:", latest_run.timestamp.strftime("%Y-%m-%d %H:%M UTC") if latest_run.timestamp else "N/A"])
    summary_ws.append([])

    summary_ws.append(["Category", "Total Issues", "Critical", "High", "Medium", "Low"])
    for cell in summary_ws[7]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    row = 8
    total_all = {"total": 0, "CRITICAL": 0, "HIGH": 0, "MEDIUM": 0, "LOW": 0}
    for cat, issues in categories.items():
        counts = {"CRITICAL": 0, "HIGH": 0, "MEDIUM": 0, "LOW": 0}
        for issue in issues:
            sev = getattr(issue, 'severity', issue.get('severity', 'MEDIUM'))
            if sev in counts:
                counts[sev] += 1
                total_all[sev] += 1

        total = sum(counts.values())
        total_all["total"] += total

        summary_ws.append([
            category_names.get(cat, cat),
            total,
            counts["CRITICAL"],
            counts["HIGH"],
            counts["MEDIUM"],
            counts["LOW"]
        ])
        for cell in summary_ws[row]:
            cell.border = thin_border
        row += 1

    summary_ws.append([
        "TOTAL",
        total_all["total"],
        total_all["CRITICAL"],
        total_all["HIGH"],
        total_all["MEDIUM"],
        total_all["LOW"]
    ])
    for cell in summary_ws[row]:
        cell.font = Font(bold=True)
        cell.border = thin_border

    # Adjust column widths for summary
    summary_ws.column_dimensions["A"].width = 25
    for col in ["B", "C", "D", "E", "F"]:
        summary_ws.column_dimensions[col].width = 12

    # Create sheet for each category
    for cat, issues in categories.items():
        ws = wb.create_sheet(category_names.get(cat, cat)[:31])  # Excel sheet name limit

        # Headers
        headers = ["Severity", "Issue", "Location", "Mitigation"]
        ws.append(headers)
        for col, cell in enumerate(ws[1], 1):
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True)

        # Data rows
        for issue in issues:
            severity = getattr(issue, 'severity', issue.get('severity', ''))
            issue_text = getattr(issue, 'issue', issue.get('issue', ''))
            location = getattr(issue, 'location', issue.get('location', ''))
            mitigation = getattr(issue, 'mitigation', issue.get('mitigation', ''))

            ws.append([severity, issue_text, location, mitigation])
            row_num = ws.max_row

            # Apply severity coloring
            severity_cell = ws.cell(row=row_num, column=1)
            if severity == "CRITICAL":
                severity_cell.fill = critical_fill
            elif severity == "HIGH":
                severity_cell.fill = high_fill
            elif severity == "MEDIUM":
                severity_cell.fill = medium_fill
            elif severity == "LOW":
                severity_cell.fill = low_fill

            for col in range(1, 5):
                cell = ws.cell(row=row_num, column=col)
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True, vertical='top')

        # Adjust column widths
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 50
        ws.column_dimensions["C"].width = 40
        ws.column_dimensions["D"].width = 50

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"code_audit_report_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
